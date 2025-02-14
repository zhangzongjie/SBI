import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Alignment
from config import SBI_USERNAME, SBI_PASSWORD, EXCEL_FILE_PATH, CHROME_USER_DATA_DIR


# 读取 Excel 文件
excel_file_path = EXCEL_FILE_PATH


    
try:
    workbook = load_workbook(excel_file_path)
except PermissionError:
    print("请关闭 Excel 文件，然后再运行此脚本。")
    exit(1)


# 设置你的证券公司网址和登录信息
url = 'https://www.sbisec.co.jp/'
username = SBI_USERNAME
password = SBI_PASSWORD

# 加载已存在的 Excel 文件
#excel_file = 'account_balance.xlsx'
#df = pd.read_excel(excel_file)

# 启动浏览器
options = webdriver.ChromeOptions()
options.add_argument(f"user-data-dir={CHROME_USER_DATA_DIR}")  # 指定用户数据目录

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get(url)

# 登录过程
print('login')
#time.sleep(2)  # 等待页面加载
driver.find_element(By.NAME, 'user_id').send_keys(username)
driver.find_element(By.NAME, 'user_password').send_keys(password + Keys.RETURN)

# 等待登录完成
#time.sleep(100)  # 根据网络情况调整等待时间

# "口座管理" 按钮
account_management_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, '//img[@alt="口座管理"]'))
)
account_management_button.click()  # 点击按钮
print('点击了口座管理按钮')  # 打印调试信息

# 点击 "My資産" 链接
account_management_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, '//img[@alt="My資産"]'))
)

# 获取当前窗口句柄
original_window = driver.current_window_handle
print(original_window)

account_management_button.click()  # 点击按钮

# 等待新窗口打开
#WebDriverWait(driver, 10).until(EC.new_window_is_opened(original_window))

# 切换到新打开的窗口
for window_handle in driver.window_handles:
    if window_handle != original_window:
        driver.switch_to.window(window_handle)
        break
#my_assets_link = WebDriverWait(driver, 10).until(
#    EC.visibility_of_element_located((By.XPATH, '//a[contains(text(), "My資産")]'))
#)
#my_assets_link.click()  # 点击链接
print('点击了My資産')  # 打印调试信息

    
account_balance_div = WebDriverWait(driver, 10).until(
    EC.visibility_of_element_located((By.XPATH, '//div[contains(@class, "css-1yowqsa")]'))
)

# 获取文本内容
account_balance = account_balance_div.text  # 获取文本内容
print(f'账户余额: {account_balance} 円')  # 打印账户余额

# 提取数字部分
account_balance_value = account_balance.split()[0]  # 提取数字部分
print(f'提取的账户余额: {account_balance_value} 円')  # 打印提取的账户余额

# 如果需要，切换回原窗口
#driver.switch_to.window(original_window)

# 记录当前日期和账户余额
current_date = datetime.now().strftime('%Y/%m/%d')  # 格式化日期为 YYYY/MM/DD
print(f'时间: {current_date} ')  # 打印时间
#df = df.append({'Date': current_date, 'Account Balance': account_balance}, ignore_index=True)

sheet = workbook.active  # 获取活动工作表


# 找到 C 列的第一个空白格
for row in range(1, sheet.max_row + 1):
    if sheet[f'C{row}'].value is None:  # 检查 C 列的单元格是否为空
        sheet[f'C{row}'] = account_balance_value  # 填入金额
        sheet[f'B{row}'] = current_date  # 填入时间
        print(f'已将账户余额 {account_balance_value} 填入 C 列的第 {row} 行。')
        print(f'已将时间 {current_date} 填入 B 列的第 {row} 行。')

        # 复制上一行的 E 和 F 列
        if row > 1:  # 确保不是第一行
            sheet[f'E{row}'] = sheet[f'E{row - 1}'].value  # 复制 E 列
            # 设置 C 列的格式为带千位分隔符的数字
            sheet[f'E{row}'].number_format = '#,##0'  # 设置格式为带千位分隔符的整数

             # 获取上一行的公式
            previous_formula = sheet[f'F{row - 1}'].value
            
            # 替换公式中的行号
            if previous_formula and isinstance(previous_formula, str):
                # 使用正则表达式替换行号
                new_formula = previous_formula.replace(f'C{row - 1}', f'C{row}')
                sheet[f'F{row}'] = new_formula  # 将调整后的公式写入当前行的 F 列
                
                # 设置 C 列的格式为带千位分隔符的数字
                sheet[f'F{row}'].number_format = '#,##0'  # 设置格式为带千位分隔符的整数

            else:
                sheet[f'F{row}'] = None  # 如果没有公式，则设置为 None
            # 清空上一行的 E 和 F 列
            sheet[f'E{row - 1}'] = None
            sheet[f'F{row - 1}'] = None
        break

# 保存修改后的 Excel 文件
workbook.save(excel_file_path)

# 关闭浏览器
driver.quit()
